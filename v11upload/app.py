from __future__ import annotations

"""
V11 of the SATS reconciliation application
-----------------------------------------

This version builds upon the V10 implementation and introduces support for
aligning SAP BFC and OneStream data at a common OS (OneStream) account level.

Key changes in V11:

* A global SAP BFC mapping workbook is used to convert SAP BFC account codes
  (including those with suffixes) into the corresponding OS account codes.
  When a GL code has an associated SAP mapping, the mapping is resolved via
  this workbook to derive the OS account.  This ensures that SAP and
  OneStream totals are compared at the same OS account level.  See
  ``data/sap_bfc_mapping.xlsx`` for the mapping definitions.

* The entity mapping CSVs (e.g. ``mapping_2403.csv``, ``mapping_2708.csv``)
  still provide the GL‑to‑OS mappings for the ERP side.  However, when a
  GL code includes a ``sap_mapping`` column, that value is now passed
  through the SAP BFC mapping workbook to derive the OS account.  The
  resulting OS account supersedes the CSV mapping, aligning both sides.

* The reconciliation is still presented in two tiers: a high‑level summary
  and a drilldown to GL codes.  The summary rows aggregate values across
  OS accounts (via the mapping described above).  For most line items the
  drilldown shows GL codes under the corresponding OS account line item.  For
  revenue (code ``4100000``), the drilldown goes directly to the GL codes.

* Debug information includes the selected entity and counts of SAP and OS
  rows, the number of mapped and unmapped GL codes, and a list of the
  largest unmapped differences.

This file contains the Flask backend.  See ``templates/index.html`` and
``static/app.js`` for the front‑end.
"""

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET
from flask import Flask, jsonify, render_template, request, send_file

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
HIERARCHY_PATH = DATA_DIR / 'hierarchy.xml'

# Per‑entity mapping files.  Each file maps ERP GL codes to OS account
# codes and optionally to SAP mappings.  The OS codes in these files may
# be overridden by the global SAP BFC mapping if a sap_mapping is provided.
ENTITY_MAPPING_FILES = {
    '2403': DATA_DIR / 'mapping_2403.csv',
    '2708': DATA_DIR / 'mapping_2708.csv',
}
DEFAULT_ENTITY = '2403'

# SAP BFC mapping workbook path.  This file maps SAP BFC account codes
# (including variants with suffixes) to OS account codes.  It is used to
# align SAP BFC data with OneStream on a common account structure.
SAP_BFC_MAPPING_PATH = DATA_DIR / 'sap_bfc_mapping.xlsx'

# ERP to SAP BFC mapping workbook path.  This file maps ERP GL codes to
# SAP BFC account codes.  It is used to supplement the per‑entity
# mapping CSVs when the ``sap_mapping`` column is missing.  See
# ``data/ERP TO SAP BFC MAPPING.xlsx`` for details.  Only used for
# certain entities (e.g. 2708) where the per‑entity mapping does not
# include a sap_mapping.
ERP_TO_SAP_BFC_MAPPING_PATH = DATA_DIR / 'ERP TO SAP BFC MAPPING.xlsx'

# Cached mapping loaded from ``SAP_BFC_MAPPING_PATH``.  Populated on
# first use by ``load_bfc_to_os_map``.
_BFC_TO_OS_MAP: dict[str, str] | None = None

app = Flask(__name__)

# Summary and drilldown definitions.  These define the high‑level line items
# and their display names.  They correspond to OS account codes in the
# hierarchy.  Revenue (4100000) is treated specially in the front end.
SUMMARY_ROWS = [
    ('4100000', 'Revenue'),
    ('EXP_EXC_DNA', 'Operating Expense (Ex-D And A)'),
    ('EBITDA', 'EBITDA'),
    ('EBIT', 'EBIT'),
    ('PBT', 'Profit Or Loss Before Tax'),
    ('PAT', 'Net Profit Or Loss (PAT)'),
    ('ATT_OWN', 'Profit Or Loss Attributable To Owners Of The Company'),
]

DRILLDOWN_ROWS = [
    ('4100000', 'Revenue'),
    ('EXP_EXC_DNA', 'Operating Expense (Ex-D And A)'),
    ('5010000', 'Cost Of Raw Materials And Supplies'),
    ('5020000', 'Staff Costs'),
    ('5030000', 'Licence Fees'),
    ('5050000', 'Company Premise Utilities And Maintenance'),
    ('5060000', 'Subcontracting services'),
    ('5080000', 'Other costs'),
    ('EBITDA', 'EBITDA'),
    ('5040000', 'Depreciation And Amortisation'),
    ('EBIT', 'EBIT'),
    ('6021000', 'Finance Income'),
    ('6031000', 'Finance Expense'),
    ('8010000', 'Share Of Results Of AJV'),
    ('6010000', 'Non operating gain loss'),
    ('6990000', 'Exceptional Items'),
    ('PBT', 'Profit Or Loss Before Tax'),
    ('6070000', 'Income Tax Expense'),
    ('8610000', 'Profit Or Loss From Discontinued Operation (Net Of Tax)'),
    ('PAT', 'Net Profit Or Loss (PAT)'),
    ('PL_MI', 'Minority Interest'),
    ('ATT_OWN', 'Profit Or Loss Attributable To Owners Of The Company'),
]

# Codes that should be highlighted in the summary and drilldown tables.
HIGHLIGHT_CODES = {'4100000', 'EXP_EXC_DNA', 'EBITDA', 'EBIT', 'PBT', 'PAT', 'ATT_OWN'}


@dataclass
class Node:
    code: str
    name: str
    parent: str | None
    level: int
    children: list[str] = field(default_factory=list)


def _stream_xlsx_rows(file_storage, header_row: int):
    stream = getattr(file_storage, 'stream', file_storage)
    stream.seek(0)
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == header_row:
            header = [str(x).strip() if x is not None else '' for x in row]
            continue
        if header is not None:
            yield header, row


def _build_index(header: list[str], required: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    missing: list[str] = []
    for name in required:
        if name not in header:
            missing.append(name)
        else:
            idx[name] = header.index(name)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return idx


def normalize_code(value: Any) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().upper().replace('–', '-').replace('—', '-').replace(' ', '').replace(',', '')
    if s.endswith('.0'):
        s = s[:-2]
    return '' if s == 'NAN' else s


def extract_gl_code(value: Any) -> str:
    s = normalize_code(value)
    if not s:
        return ''
    m = re.match(r'^([A-Z0-9_]+)', s)
    return m.group(1) if m else s


def first_nonblank(series: pd.Series) -> str:
    for v in series:
        if pd.notna(v):
            s = str(v).strip()
            if s and s.upper() != 'NAN':
                return s
    return ''


def parse_hierarchy(path: Path) -> dict[str, Node]:
    tree = ET.parse(path)
    root = tree.getroot()
    nodes: dict[str, Node] = {}

    def walk(elem: ET.Element, parent: str | None, level: int):
        code = normalize_code(elem.attrib.get('code'))
        name = (elem.attrib.get('name') or '').strip()
        if not code:
            return
        nodes[code] = Node(code=code, name=name, parent=parent, level=level)
        if parent and parent in nodes:
            nodes[parent].children.append(code)
        for child in elem.findall('./Account'):
            walk(child, code, level + 1)

    for acct in root.findall('./Account'):
        walk(acct, None, 0)
    return nodes


def ancestors_of(code: str, nodes: dict[str, Node]) -> set[str]:
    out: set[str] = set()
    cur = code
    seen = set()
    while cur and cur not in seen and cur in nodes:
        out.add(cur)
        seen.add(cur)
        cur = nodes[cur].parent or ''
    return out


def load_bfc_to_os_map() -> dict[str, str]:
    """Load the SAP BFC to OS account mapping as a dictionary.

    Returns a mapping from normalised SAP BFC account codes (both original
    and cleaned) to normalised OS account codes.  The result is cached
    globally to avoid repeated file reads.
    """
    global _BFC_TO_OS_MAP
    if _BFC_TO_OS_MAP is not None:
        return _BFC_TO_OS_MAP
    mapping: dict[str, str] = {}
    try:
        df = pd.read_excel(SAP_BFC_MAPPING_PATH)
    except Exception:
        _BFC_TO_OS_MAP = {}
        return _BFC_TO_OS_MAP
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    bfc_cols = [c for c in df.columns if 'SAP BFC' in c.upper()]
    conv_cols = [c for c in df.columns if 'CONVERT' in c.upper()]
    os_cols = [c for c in df.columns if 'BFC TO OS' in c.upper()]
    if not bfc_cols or not os_cols:
        _BFC_TO_OS_MAP = {}
        return _BFC_TO_OS_MAP
    bfc_col = bfc_cols[0]
    conv_col = conv_cols[0] if conv_cols else None
    os_col = os_cols[0]
    for _, row in df.iterrows():
        os_code = normalize_code(row.get(os_col, ''))
        if not os_code:
            continue
        orig = normalize_code(row.get(bfc_col, ''))
        if orig:
            mapping[orig] = os_code
        if conv_col:
            conv = normalize_code(row.get(conv_col, ''))
            if conv:
                mapping[conv] = os_code
    _BFC_TO_OS_MAP = mapping
    return mapping


def load_entity_mapping(entity: str, nodes: dict[str, Node]) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """Load the mapping for a specific entity.

    Each mapping CSV must contain the columns: gl_code, os_leaf_code,
    sap_mapping, gl_name, sap_description.  In V11, when a sap_mapping
    value is present, it is resolved through the global SAP BFC mapping
    workbook to derive an OS account code.  That OS code supersedes the
    os_leaf_code from the CSV.  The resulting data frame includes a
    ``line_items`` column derived from the account hierarchy.

    Args:
        entity: The entity code (e.g. '2403', '2708').
        nodes: The account hierarchy as returned by ``parse_hierarchy``.

    Returns:
        A tuple of (meta, mapping_rows) where meta is a DataFrame with
        columns gl_code, os_leaf_code, sap_mapping, gl_name, sap_description,
        line_items.  mapping_rows is a list of dicts for JSON output.
    """
    mapping_path = ENTITY_MAPPING_FILES.get(entity)
    if not mapping_path or not mapping_path.exists():
        raise ValueError(f"No mapping file defined for entity {entity}")
    meta = pd.read_csv(mapping_path).copy()
    # Normalise key columns
    def _norm_series(s):
        return s.fillna('').astype(str).map(normalize_code)
    for col in ['gl_code', 'os_leaf_code', 'sap_mapping']:
        if col in meta.columns:
            meta[col] = _norm_series(meta[col])
        else:
            meta[col] = ''
    # Ensure descriptive columns exist
    for col in ['gl_name', 'sap_description']:
        if col not in meta.columns:
            meta[col] = ''
        else:
            meta[col] = meta[col].fillna('').astype(str)
    # Load the global BFC mapping
    try:
        bfc_map = load_bfc_to_os_map()
    except Exception:
        bfc_map = {}
    # If this entity has no sap_mapping values in the CSV (e.g. 2708),
    # attempt to supplement them from the ERP→SAP BFC mapping workbook.
    # This allows SAP BFC raw data to be linked back to GL codes via
    # the mapping file without manual coding.  Only populate missing
    # sap_mapping entries so that any explicit values in the CSV take
    # precedence.
    if entity and entity.strip().startswith('2708'):
        try:
            erp_bfc_df = pd.read_excel(ERP_TO_SAP_BFC_MAPPING_PATH)
            # normalise column names for consistency
            erp_bfc_df.columns = [str(c).strip() for c in erp_bfc_df.columns]
            # expect columns 'LOCAL ACC' and 'D_A'
            if 'LOCAL ACC' in erp_bfc_df.columns and 'D_A' in erp_bfc_df.columns:
                # build lookup dict for gl_code -> sap_mapping
                erp_bfc_df['LOCAL ACC'] = erp_bfc_df['LOCAL ACC'].astype(str).map(normalize_code)
                erp_bfc_df['D_A'] = erp_bfc_df['D_A'].astype(str).map(normalize_code)
                lookup_map = erp_bfc_df.set_index('LOCAL ACC')['D_A'].to_dict()
                # fill missing sap_mapping values using the lookup
                def _fill_sap_mapping(row):
                    if row.get('sap_mapping', ''):
                        return row['sap_mapping']
                    gl = normalize_code(row.get('gl_code', ''))
                    return lookup_map.get(gl, '')
                meta['sap_mapping'] = meta.apply(_fill_sap_mapping, axis=1)
        except Exception:
            # ignore any errors reading the ERP mapping; fall back to CSV values
            pass

    # Resolve OS account code: prefer SAP mapping via BFC mapping, otherwise
    # use the CSV mapping.  Normalise codes to ensure consistency.
    def _resolve_os_code(row):
        sap_map = row.get('sap_mapping', '')
        sap_map_norm = normalize_code(sap_map)
        if sap_map_norm and sap_map_norm in bfc_map:
            return bfc_map[sap_map_norm]
        return normalize_code(row.get('os_leaf_code', ''))
    meta['os_leaf_code'] = meta.apply(_resolve_os_code, axis=1)
    # Derive line items from the OS code using the hierarchy.  If no OS
    # code is available, leave the line_items empty.  Sorting ensures
    # deterministic ordering in the UI.
    meta['line_items'] = meta['os_leaf_code'].apply(
        lambda x: sorted(ancestors_of(x, nodes)) if isinstance(x, str) and x else []
    )
    # Compose mapping rows for UI display
    mapping_rows = meta[['gl_code', 'os_leaf_code', 'sap_mapping', 'gl_name', 'sap_description']].fillna('').to_dict(orient='records')
    return meta, mapping_rows


def _read_excel_all_sheets(file_storage) -> dict[str, pd.DataFrame]:
    stream = getattr(file_storage, 'stream', file_storage)
    stream.seek(0)
    return pd.read_excel(stream, sheet_name=None, header=None)


def _find_header_row(df: pd.DataFrame, required: list[str], aliases: dict[str, list[str]] | None = None, scan_rows: int = 20) -> tuple[int, list[str]]:
    aliases = aliases or {}
    required_sets = {name: {name, *aliases.get(name, [])} for name in required}
    limit = min(scan_rows, len(df))
    for i in range(limit):
        row_vals = [str(x).strip() if pd.notna(x) else '' for x in df.iloc[i].tolist()]
        row_set = set(row_vals)
        if all(any(alias in row_set for alias in required_sets[name]) for name in required):
            return i, row_vals
    raise ValueError(f"Could not find a header row containing: {', '.join(required)}")


def _canonicalize_columns(df: pd.DataFrame, aliases: dict[str, list[str]]) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        c = str(col).strip()
        for canonical, opts in aliases.items():
            if c == canonical or c in opts:
                rename[col] = canonical
                break
    return df.rename(columns=rename)


def read_sap(file_storage) -> pd.DataFrame:
    """Read the SAP BFC workbook and return a DataFrame of relevant columns.

    V11 supports two formats:
    1. The new TB BFC format used in the 2708 dataset (columns LOCAL ACC and P_AMOUNT).
    2. The original ERP→SAP BFC workbook (GL Code, SAP Mapping, Amount, etc.).

    The returned DataFrame always contains the columns: GL Code, GL Name,
    SAP Mapping, SAP Description, Amount, along with placeholder columns
    for Month, P&L Nos, P&L Head, CC Code, Cost Center Name and MIS Type.
    """
    required = ['Month', 'P&L Nos', 'P&L Head', 'GL Code', 'GL Name', 'CC Code', 'Cost Center Name', 'MIS Type', 'SAP Mapping', 'SAP Description', 'Amount']
    aliases = {
        'GL Code': ['GL code', 'GLCode'],
        'GL Name': ['GL name', 'GLName'],
        'SAP Mapping': ['SAP mapping', 'SAPMapping'],
        'SAP Description': ['SAP description', 'SAPDescription'],
    }
    sheets = _read_excel_all_sheets(file_storage)
    # Attempt to parse TB BFC first
    for sheet_name, raw in sheets.items():
        header_candidate = [str(x).strip() if pd.notna(x) else '' for x in raw.iloc[0].tolist()]
        if 'LOCAL ACC' in header_candidate and 'P_AMOUNT' in header_candidate:
            df = raw.copy()
            df.columns = header_candidate
            if 'LOCAL ACC' not in df.columns or 'P_AMOUNT' not in df.columns:
                continue
            out = pd.DataFrame()
            out['GL Code'] = df['LOCAL ACC'].map(lambda v: extract_gl_code(v))
            out['GL Name'] = ''
            out['SAP Mapping'] = ''
            out['SAP Description'] = ''
            out['Amount'] = pd.to_numeric(df['P_AMOUNT'], errors='coerce').fillna(0.0)
            out['Month'] = ''
            out['P&L Nos'] = ''
            out['P&L Head'] = ''
            out['CC Code'] = ''
            out['Cost Center Name'] = ''
            out['MIS Type'] = ''
            out = out[out['GL Code'].astype(str).str.strip() != ''].copy().reset_index(drop=True)
            return out.reset_index(drop=True)
    # Fall back to ERP→SAP BFC format
    picked = None
    for sheet_name, raw in sheets.items():
        try:
            header_idx, header = _find_header_row(raw, ['GL Code', 'SAP Mapping', 'Amount'], aliases=aliases)
            picked = (sheet_name, raw, header_idx, header)
            break
        except Exception:
            continue
    if picked is None:
        raise ValueError('Unable to find a valid header row in the ERP to SAP BFC workbook.')
    _, raw, header_idx, header = picked
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = header
    df = _canonicalize_columns(df, aliases)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ERP to SAP BFC workbook is missing required columns: {', '.join(missing)}")
    df = df[required].copy()
    df = df[~df.apply(lambda r: all(v is None or str(v).strip() == '' for v in r), axis=1)].copy()
    return df.reset_index(drop=True)


def read_os(file_storage) -> pd.DataFrame:
    """Read the OneStream workbook and return a DataFrame.

    V11 relaxes the requirements on the OS workbook.  It looks for
    headers matching ``Local COA`` and ``Amount`` (case insensitive) and
    allows the columns ``OS COA`` and ``Function`` to be missing.  Any
    additional columns are preserved but ignored.  If neither of the
    required columns are found, an error is raised.
    """
    sheets = _read_excel_all_sheets(file_storage)
    aliases = {
        'OS COA': ['OS Account', 'OS account', 'OS COA '],
        'Local COA': ['GL Code', 'LocalCOA', 'Local COA '],
        'Function': ['P&L Head', 'Function '],
        'Amount': ['Amount ', 'Amt'],
    }
    picked = None
    for _, raw in sheets.items():
        try:
            header_idx, header = _find_header_row(raw, ['Local COA', 'Amount'], aliases=aliases)
            picked = (raw, header_idx, header)
            break
        except Exception:
            continue
    if picked is None:
        raise ValueError('Unable to find a valid header row in the ERP to OS workbook.')
    raw, header_idx, header = picked
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = header
    df = _canonicalize_columns(df, aliases)
    df.columns = [str(c).strip() for c in df.columns]
    if 'OS COA' not in df.columns and 'OS Account' in df.columns:
        df = df.rename(columns={'OS Account': 'OS COA'})
    if 'Function' not in df.columns:
        df['Function'] = ''
    if 'OS COA' not in df.columns:
        df['OS COA'] = ''
    required = ['Amount', 'Local COA']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"ERP to OS workbook is missing required columns: {', '.join(missing)}")
    df = df[~df.apply(lambda r: all(v is None or str(v).strip() == '' for v in r), axis=1)].copy()
    return df.reset_index(drop=True)


def build_gl_totals(sap: pd.DataFrame, os_df: pd.DataFrame, meta: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Aggregate SAP and OS data at the GL code level.

    This function merges SAP and OS data on GL codes, attaches descriptive
    information from the mapping, and computes differences.  The resulting
    DataFrame contains the columns gl_code, os_leaf_code, sap_bfc,
    onestream, difference, description, currency, and line_items.  A
    debug dictionary includes summary statistics about the reconciliation.

    The function relies on ``meta`` having its ``os_leaf_code`` values
    already resolved via the global SAP BFC mapping, as performed in
    ``load_entity_mapping``.
    """
    sap = sap.copy()
    sap['gl_code_raw'] = sap['GL Code'].map(extract_gl_code)
    sap['sap_mapping_clean'] = sap['SAP Mapping'].map(normalize_code)
    sap['amount_display'] = -pd.to_numeric(sap['Amount'], errors='coerce').fillna(0.0) / 1000.0

    # Use sap_mapping to fill missing gl_code if possible.  This behaviour
    # comes from V10 and is retained for backward compatibility.  If a
    # mapping exists from sap_mapping_clean to a gl_code in the meta, fill it.
    sap_map_lookup = meta[['gl_code', 'sap_mapping']].dropna().copy()
    sap_map_lookup['sap_mapping'] = sap_map_lookup['sap_mapping'].map(normalize_code)
    sap_map_lookup['gl_code'] = sap_map_lookup['gl_code'].map(normalize_code)
    sap_map_lookup = sap_map_lookup[sap_map_lookup['sap_mapping'] != ''].drop_duplicates('sap_mapping')
    sap = sap.merge(
        sap_map_lookup.rename(columns={'gl_code': 'gl_code_from_mapping'}),
        left_on='sap_mapping_clean',
        right_on='sap_mapping',
        how='left'
    )
    sap['gl_code'] = sap['gl_code_raw']
    sap.loc[sap['gl_code'].eq('') & sap['gl_code_from_mapping'].notna(), 'gl_code'] = sap['gl_code_from_mapping']

    # Aggregate SAP by gl_code
    sap_grp = sap[sap['gl_code'] != ''].groupby('gl_code', dropna=False).agg(
        sap_bfc=('amount_display', 'sum'),
        sap_description=('GL Name', first_nonblank),
        sap_currency=('Month', lambda s: ''),
    ).reset_index()

    # Aggregate OS by gl_code
    os_df = os_df.copy()
    os_df['gl_code'] = os_df['Local COA'].map(extract_gl_code)
    os_df['amount_display'] = -pd.to_numeric(os_df['Amount'], errors='coerce').fillna(0.0) / 1000.0
    currency_col = next((c for c in ['Currency', 'Curr', 'Local Currency', 'CCY'] if c in os_df.columns), None)
    if currency_col:
        os_df['currency'] = os_df[currency_col].fillna('').astype(str).str.strip()
    else:
        os_df['currency'] = ''
    os_grp = os_df[os_df['gl_code'] != ''].groupby('gl_code', dropna=False).agg(
        onestream=('amount_display', 'sum'),
        os_currency=('currency', first_nonblank),
        function=('Function', first_nonblank),
    ).reset_index()

    # Merge SAP and OS totals with the mapping
    all_gl = meta.merge(sap_grp, on='gl_code', how='outer').merge(os_grp, on='gl_code', how='outer')
    all_gl['sap_bfc'] = pd.to_numeric(all_gl.get('sap_bfc'), errors='coerce').fillna(0.0)
    all_gl['onestream'] = pd.to_numeric(all_gl.get('onestream'), errors='coerce').fillna(0.0)
    all_gl['difference'] = all_gl['sap_bfc'] - all_gl['onestream']
    # Prefer GL name for description; fallback to OS description (not present in CSV) or blank
    all_gl['description'] = all_gl['gl_name'].where(all_gl.get('gl_name', '').fillna('').astype(str).str.strip() != '', '')
    all_gl['currency'] = all_gl.get('os_currency', '').fillna('')
    all_gl['line_items'] = all_gl['line_items'].apply(lambda x: x if isinstance(x, list) else [])
    all_gl['gl_code'] = all_gl['gl_code'].fillna('').astype(str)
    all_gl = all_gl[all_gl['gl_code'].str.strip() != ''].copy()

    debug = {
        'sap_rows': int(len(sap)),
        'os_rows': int(len(os_df)),
        'sap_gl_mapped': int(sap_grp['gl_code'].nunique()),
        'os_gl_mapped': int(os_grp['gl_code'].nunique()),
        'all_gl_codes': int(all_gl['gl_code'].nunique()),
        'os_rows_with_amount': int((pd.to_numeric(os_df['Amount'], errors='coerce').fillna(0.0) != 0).sum()),
        'os_gl_codes_matched_to_mapping': int(all_gl[(all_gl['onestream'] != 0) & (all_gl['line_items'].apply(len) > 0)]['gl_code'].nunique()),
        'os_gl_codes_unmapped': int(all_gl[(all_gl['onestream'] != 0) & (all_gl['line_items'].apply(len) == 0)]['gl_code'].nunique()),
    }
    return all_gl, debug


def line_item_gls(all_gl: pd.DataFrame, line_code: str) -> pd.DataFrame:
    """Return the subset of GL totals that roll up to the given line code."""
    return all_gl[all_gl['line_items'].apply(lambda xs: line_code in xs)].copy()


def build_summary_rows(all_gl: pd.DataFrame) -> list[dict[str, Any]]:
    rows = []
    for code, name in SUMMARY_ROWS:
        part = line_item_gls(all_gl, code)
        sap = round(float(part['sap_bfc'].sum()), 2)
        os = round(float(part['onestream'].sum()), 2)
        rows.append({
            'code': code,
            'name': name,
            'sap_bfc': sap,
            'onestream': os,
            'difference': round(sap - os, 2),
            'highlight': code in HIGHLIGHT_CODES,
        })
    return rows


def build_drilldown_rows(all_gl: pd.DataFrame) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for code, name in DRILLDOWN_ROWS:
        part = line_item_gls(all_gl, code).sort_values(['difference', 'sap_bfc', 'onestream'], key=lambda s: s.abs(), ascending=False)
        sap = round(float(part['sap_bfc'].sum()), 2)
        os = round(float(part['onestream'].sum()), 2)
        out.append({
            'row_type': 'parent',
            'code': code,
            'name': name,
            'description': '',
            'currency': '',
            'sap_bfc': sap,
            'onestream': os,
            'difference': round(sap - os, 2),
            'highlight': code in HIGHLIGHT_CODES,
            'child_count': int(len(part)),
        })
        for _, r in part.iterrows():
            out.append({
                'row_type': 'child',
                'parent_code': code,
                'code': r['gl_code'],
                'name': r['gl_code'],
                'description': r.get('description', '') or '',
                'currency': r.get('currency', '') or '',
                'sap_bfc': round(float(r['sap_bfc']), 2),
                'onestream': round(float(r['onestream']), 2),
                'difference': round(float(r['difference']), 2),
                'highlight': False,
            })
    return out


def process_files(sap_file, os_file, entity: str = DEFAULT_ENTITY) -> dict[str, Any]:
    """Run reconciliation for the given SAP and OS files, using the mapping for the specified entity."""
    nodes = parse_hierarchy(HIERARCHY_PATH)
    meta, mapping_rows = load_entity_mapping(entity, nodes)
    sap = read_sap(sap_file)
    os_df = read_os(os_file)
    all_gl, debug = build_gl_totals(sap, os_df, meta)
    summary_rows = build_summary_rows(all_gl)
    drilldown_rows = build_drilldown_rows(all_gl)
    # Determine unmapped GL codes (no line items)
    unmapped_gls = all_gl[all_gl['line_items'].apply(len) == 0].sort_values('difference', key=lambda s: s.abs(), ascending=False).head(200)
    unmapped_rows = [
        {
            'gl_code': r['gl_code'],
            'description': r.get('description', '') or '',
            'sap_bfc': round(float(r['sap_bfc']), 2),
            'onestream': round(float(r['onestream']), 2),
            'difference': round(float(r['difference']), 2),
        }
        for _, r in unmapped_gls.iterrows()
    ]
    debug.update({
        'sap_total_all_rows': round(float(all_gl['sap_bfc'].sum()), 2),
        'os_total_all_rows': round(float(all_gl['onestream'].sum()), 2),
        'unmapped_gl_codes': int((all_gl['line_items'].apply(len) == 0).sum()),
        'unmapped_top_items': unmapped_rows,
        'entity': entity,
    })
    return {
        'summary_rows': summary_rows,
        'drilldown_rows': drilldown_rows,
        'debug': debug,
        'mapping_rows': mapping_rows[:1000],
    }


def build_export_workbook(results: dict[str, Any]) -> bytes:
    """Build the reconciliation workbook for download as an Excel file."""
    output = io.BytesIO()
    summary_df = pd.DataFrame(results['summary_rows'])
    drill_df = pd.DataFrame(results['drilldown_rows'])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        drill_df.to_excel(writer, index=False, sheet_name='Drilldown')
        drill_df[(drill_df['row_type'] == 'parent') | (drill_df['sap_bfc'] != 0) | (drill_df['onestream'] != 0)].to_excel(writer, index=False, sheet_name='Drilldown(no zeroes)')
        wb = writer.book
        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    value = '' if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                    if ws.title.startswith('Drilldown') and cell.row > 1:
                        code_value = ws.cell(cell.row, 2).value
                        if code_value in HIGHLIGHT_CODES:
                            cell.font = openpyxl.styles.Font(bold=True)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)
    output.seek(0)
    return output.getvalue()


@app.get('/')
def index():
    return render_template('index.html')


@app.post('/api/run-recon')
def run_recon():
    sap_file = request.files.get('sap_file')
    os_file = request.files.get('os_file')
    if not sap_file or not os_file:
        return jsonify({'error': 'Please upload both SAP BFC and OneStream files.'}), 400
    entity = request.form.get('entity', DEFAULT_ENTITY)
    try:
        return jsonify(process_files(sap_file, os_file, entity))
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


@app.post('/api/export')
def export_results():
    sap_file = request.files.get('sap_file')
    os_file = request.files.get('os_file')
    if not sap_file or not os_file:
        return jsonify({'error': 'Please upload both SAP BFC and OneStream files.'}), 400
    entity = request.form.get('entity', DEFAULT_ENTITY)
    try:
        content = build_export_workbook(process_files(sap_file, os_file, entity))
        return send_file(
            io.BytesIO(content),
            as_attachment=True,
            download_name='reconciliation_output_v11.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400


if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)